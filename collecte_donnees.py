# -*- coding: utf-8 -*-

import os
import requests
import time
import logging
import argparse
from bs4 import BeautifulSoup
from datetime import datetime
from io import BytesIO
from PIL import Image
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

# Configurer le logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s [%(levelname)s] %(message)s',
    handlers=[
        logging.FileHandler("scraping.log"),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

def load_previous_products(output_file):
    """Charge les liens des produits du fichier Excel précédent."""
    previous_links = set()
    if os.path.exists(output_file):
        try:
            wb = load_workbook(output_file)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):  # Commence à la deuxième ligne pour ignorer les en-têtes
                lien = row[-1]  # Le lien est supposé être dans la dernière colonne
                if lien:
                    previous_links.add(lien)
            logger.info(f"{len(previous_links)} produits chargés depuis le fichier précédent.")
        except Exception as e:
            logger.error(f"Erreur lors du chargement du fichier Excel précédent : {e}")
    else:
        logger.info("Aucun fichier Excel précédent trouvé.")
    return previous_links

def collect_products(max_products=60, url='https://www.aliexpress.com/p/calp-plus/index.html', scroll_pause_time=2.0, previous_links=set()):
    # Configurer Selenium pour Chrome en mode sans tête
    options = Options()
    options.headless = True  # Mode headless pour ne pas ouvrir de fenêtre de navigateur

    # Initialiser la liste des produits et un ensemble pour les liens uniques
    liste_produits = []
    liens_uniques = previous_links.copy()  # Inclure les liens des produits précédemment collectés

    try:
        # Configurer Selenium pour utiliser ChromeDriver
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.set_window_size(1920, 1080)  # Définir la taille de la fenêtre pour s'assurer que tous les éléments sont visibles

        # Accéder à l'URL cible
        driver.get(url)

        # Attendre que la page soit complètement chargée
        time.sleep(5)

        produits_collectes = 0
        last_products_count = 0
        max_scrolls = 100  # Limite pour éviter une boucle infinie
        scrolls = 0

        while produits_collectes < max_products and scrolls < max_scrolls:
            # Descendre jusqu'en bas de la page
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(scroll_pause_time)

            # Récupérer les éléments de produit
            produits_elements = driver.find_elements(By.CSS_SELECTOR, "a._1UZxx")
            current_products_count = len(produits_elements)

            # Vérifier si de nouveaux produits ont été chargés
            if current_products_count == last_products_count:
                logger.info("Aucun nouveau produit n'a été chargé.")
                break

            # Traiter les nouveaux produits
            for produit_element in produits_elements[last_products_count:]:
                produit_html = produit_element.get_attribute('outerHTML')
                produit_soup = BeautifulSoup(produit_html, 'html.parser')

                # Lien du produit
                try:
                    lien = produit_element.get_attribute('href')
                    if not lien:
                        continue
                    if lien.startswith('//'):
                        lien = 'https:' + lien
                    elif lien.startswith('/'):
                        lien = 'https://www.aliexpress.com' + lien
                    if lien in liens_uniques:
                        continue  # Ignorer les produits déjà collectés
                    liens_uniques.add(lien)
                except Exception as e:
                    logger.error(f"Erreur lors de la récupération du lien du produit: {e}")
                    continue

                # Nom du produit
                try:
                    nom = produit_soup.find('h3', class_='nXeOv').get_text(strip=True)
                except AttributeError:
                    nom = 'N/A'

                # Prix du produit
                try:
                    prix_elements = produit_soup.find('div', class_='U-S0j').find_all('span')
                    prix = ''.join([elem.get_text() for elem in prix_elements]).replace('€', '').strip()
                except AttributeError:
                    prix = 'N/A'

                # Image du produit
                try:
                    image_tag = produit_soup.find('img', class_='_1IH3l product-img')
                    if image_tag:
                        image_url = image_tag.get('src') or image_tag.get('image-src')
                        if image_url.startswith('//'):
                            image_url = 'https:' + image_url
                    else:
                        image_url = 'N/A'
                except (AttributeError, TypeError):
                    image_url = 'N/A'

                # Métadonnées supplémentaires
                # Nombre de ventes
                try:
                    ventes = produit_soup.find('span', class_='jmSdc').get_text(strip=True)
                except AttributeError:
                    ventes = 'N/A'

                # Évaluations (notes)
                try:
                    evaluations = produit_soup.find('span', class_='eXPaM').get_text(strip=True)
                except AttributeError:
                    evaluations = 'N/A'

                # Nombre d'avis clients
                try:
                    avis = produit_soup.find('span', class_='ZwoRt').get_text(strip=True)
                except AttributeError:
                    avis = 'N/A'

                # Nom du vendeur
                try:
                    vendeur = produit_soup.find('a', class_='ox0KZ').get_text(strip=True)
                except AttributeError:
                    vendeur = 'N/A'

                # Lieu d'expédition
                try:
                    lieu = produit_soup.find('span', class_='Rm8mX').get_text(strip=True)
                except AttributeError:
                    lieu = 'N/A'

                # Créer le dictionnaire du produit
                dict_produit = {
                    'Nom du Produit': nom,
                    'Prix': prix,
                    'Ventes': ventes,
                    'Évaluations': evaluations,
                    'Avis': avis,
                    'Vendeur': vendeur,
                    'Lieu': lieu,
                    'Image': image_url,
                    'Lien': lien
                }

                # Ajouter le produit à la liste
                liste_produits.append(dict_produit)
                produits_collectes += 1

                if produits_collectes >= max_products:
                    logger.info(f"{max_products} nouveaux produits collectés. Fin de la collecte.")
                    break

            last_products_count = current_products_count
            scrolls += 1

            if produits_collectes >= max_products:
                break

        logger.info(f"Nombre total de nouveaux produits collectés : {produits_collectes}")

    except Exception as e:
        logger.error(f"Erreur lors de la collecte des données : {e}")
    finally:
        # Fermer le navigateur après avoir récupéré le contenu
        driver.quit()

    return liste_produits

def save_to_excel(liste_produits, output_file):
    # Vérifier si le fichier Excel existe déjà
    if os.path.exists(output_file):
        # Charger le fichier existant
        wb = load_workbook(output_file)
        ws = wb.active
        start_row = ws.max_row + 1  # Commencer à la première ligne vide
        logger.info(f"Ajout des nouveaux produits à la suite du fichier existant.")
    else:
        # Créer un nouveau fichier Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Produits Viraux"
        start_row = 2  # Commencer à la deuxième ligne (la première est pour les en-têtes)

        # Ajouter les en-têtes avec les nouvelles métadonnées
        ws.append(['Nom du Produit', 'Prix', 'Ventes', 'Évaluations', 'Avis', 'Vendeur', 'Lieu', 'Image', 'Lien'])
        logger.info("En-têtes ajoutées")

    # Déterminer le dossier pour enregistrer les images
    directory_name = os.path.dirname(output_file)
    if not directory_name:
        directory_name = '.'  # Dossier courant
    images_directory = os.path.join(directory_name, 'images')
    os.makedirs(images_directory, exist_ok=True)

    # Ajouter les données et les images
    for index, produit in enumerate(liste_produits, start=start_row):
        logger.info(f"Ajout du produit : {produit['Nom du Produit']}")
        ws.cell(row=index, column=1, value=produit['Nom du Produit'])
        ws.cell(row=index, column=2, value=produit['Prix'])
        ws.cell(row=index, column=3, value=produit['Ventes'])
        ws.cell(row=index, column=4, value=produit['Évaluations'])
        ws.cell(row=index, column=5, value=produit['Avis'])
        ws.cell(row=index, column=6, value=produit['Vendeur'])
        ws.cell(row=index, column=7, value=produit['Lieu'])
        ws.cell(row=index, column=9, value=produit['Lien'])

        # Télécharger et ajouter l'image en colonne 8 ('H')
        if produit['Image'] != 'N/A':
            try:
                image_response = requests.get(produit['Image'], timeout=10)
                if image_response.status_code == 200:
                    img_data = BytesIO(image_response.content)
                    img = Image.open(img_data)
                    img.thumbnail((100, 100))  # Redimensionner l'image pour qu'elle tienne dans la cellule

                    # Sauvegarder l'image dans le dossier des images
                    img_path = os.path.join(images_directory, f"image_{index}.png")
                    img.save(img_path)

                    # Insérer l'image dans Excel
                    excel_img = ExcelImage(img_path)
                    excel_img.anchor = f'H{index}'
                    ws.add_image(excel_img)
                    logger.info(f"Image ajoutée pour {produit['Nom du Produit']}")
                else:
                    logger.warning(f"Échec du téléchargement de l'image pour {produit['Nom du Produit']}")
                    ws.cell(row=index, column=8, value='N/A')
            except Exception as e:
                logger.error(f"Impossible de télécharger l'image pour {produit['Nom du Produit']}: {e}")
                ws.cell(row=index, column=8, value='N/A')
        else:
            ws.cell(row=index, column=8, value='N/A')

    # Enregistrer le fichier Excel
    wb.save(output_file)
    logger.info(f"Fichier Excel mis à jour : {output_file}")

def main():
    # Ajout des options de ligne de commande
    parser = argparse.ArgumentParser(
        description='Script pour collecter des produits depuis AliExpress et les enregistrer dans un fichier Excel.',
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    parser.add_argument('--max-products', type=int, default=60, help='Nombre maximum de nouveaux produits à collecter.')
    parser.add_argument('--url', type=str, default='https://www.aliexpress.com/p/calp-plus/index.html', help='URL de la page à scraper.')
    parser.add_argument('--output', type=str, default='produits_viraux.xlsx', help='Chemin du fichier Excel de sortie.')
    parser.add_argument('--scroll-pause', type=float, default=2.0, help='Temps d\'attente (en secondes) après chaque défilement.')
    args = parser.parse_args()

    # Charger les produits précédemment collectés
    previous_links = load_previous_products(args.output)

    # Appeler la fonction de collecte avec les arguments fournis
    liste_produits = collect_products(
        max_products=args.max_products,
        url=args.url,
        scroll_pause_time=args.scroll_pause,
        previous_links=previous_links
    )

    if liste_produits:
        save_to_excel(liste_produits, output_file=args.output)
    else:
        logger.warning("Aucun nouveau produit n'a été collecté.")

if __name__ == "__main__":
    main()
