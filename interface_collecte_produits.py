#-*- coding: utf-8 -*-

import tkinter as tk
from tkinter import messagebox, scrolledtext
from tkinter import PhotoImage
import os
import requests
from bs4 import BeautifulSoup
import time
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from PIL import Image, ImageTk
from io import BytesIO
from datetime import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

# Fonction qui exécute le script de collecte des produits
def collecter_produits():
    try:
        update_log("Démarrage de la collecte des produits...")
        
        options = Options()
        options.headless = True  # Mode headless pour ne pas ouvrir de fenêtre de navigateur

        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        url = 'https://www.aliexpress.com/p/calp-plus/index.html'
        driver.get(url)
        update_log("Ouverture de la page AliExpress...")

        time.sleep(5)

        scroll_pause_time = 2
        last_height = driver.execute_script("return document.body.scrollHeight")

        produits_collectes = 0
        max_produits = 60

        liste_produits = []

        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            update_log("Défilement vers le bas de la page...")

            time.sleep(scroll_pause_time)

            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                update_log("Fin du défilement, aucun nouveau produit à charger.")
                break

            last_height = new_height

            page_source = driver.page_source
            soup = BeautifulSoup(page_source, 'html.parser')
            produits = soup.find_all('a', class_='_1UZxx')

            for produit in produits:
                try:
                    nom = produit.find('h3', class_='nXeOv').text.strip()
                except AttributeError:
                    nom = 'N/A'

                try:
                    prix_elements = produit.find('div', class_='U-S0j').find_all('span')
                    prix = ''.join([elem.text for elem in prix_elements]).replace('€', '').strip()
                except AttributeError:
                    prix = 'N/A'

                try:
                    image_url = produit.find('img', class_='_1IH3l product-img')['src']
                    if image_url.startswith('//'):
                        image_url = 'https:' + image_url
                except (AttributeError, TypeError):
                    image_url = 'N/A'

                try:
                    lien = produit['href']
                    if lien.startswith('//'):
                        lien = 'https:' + lien
                    else:
                        lien = 'https://www.aliexpress.com' + lien
                except KeyError:
                    lien = 'N/A'

                dict_produit = {
                    'Nom du Produit': nom,
                    'Prix': prix,
                    'Image': image_url,
                    'Lien': lien
                }

                liste_produits.append(dict_produit)
                produits_collectes += 1

                if produits_collectes >= max_produits:
                    update_log(f"{max_produits} produits collectés. Fin de la collecte.")
                    break

            if produits_collectes >= max_produits:
                break

        driver.quit()
        update_log("Nouveaux produits récupérés, fermeture du navigateur.")

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        directory_name = f'produits_{timestamp}'
        os.makedirs(directory_name)
        update_log(f"Création du dossier : {directory_name}")

        file_name = f'{directory_name}/produits_viraux_{timestamp}.xlsx'

        wb = Workbook()
        ws = wb.active
        ws.title = "Produits Viraux"

        ws.append(['Nom du Produit', 'Prix', 'Image', 'Lien'])
        update_log("Ajout des en-têtes dans le fichier Excel.")

        for index, produit in enumerate(liste_produits, start=2):
            ws.cell(row=index, column=1, value=produit['Nom du Produit'])
            ws.cell(row=index, column=2, value=produit['Prix'])
            ws.cell(row=index, column=4, value=produit['Lien'])

            if produit['Image'] != 'N/A':
                try:
                    image_response = requests.get(produit['Image'])
                    img_data = BytesIO(image_response.content)
                    img = Image.open(img_data)
                    img.thumbnail((100, 100))

                    img_path = f"{directory_name}/image_{index}.png"
                    img.save(img_path)

                    excel_img = ExcelImage(img_path)
                    ws.add_image(excel_img, f'C{index}')
                except Exception as e:
                    update_log(f"Impossible de télécharger l'image pour {produit['Nom du Produit']}: {e}")
            else:
                ws.cell(row=index, column=3, value='N/A')

        wb.save(file_name)
        update_log(f"Fichier Excel enregistré : {file_name}")

        messagebox.showinfo("Succès", f"Les produits ont été collectés et sauvegardés dans le fichier : {file_name}")
    except Exception as e:
        update_log(f"Erreur : {e}")
        messagebox.showerror("Erreur", f"Une erreur s'est produite : {e}")

# Fonction pour mettre à jour les logs en temps réel
def update_log(message):
    log_text.config(state=tk.NORMAL)
    log_text.insert(tk.END, f"{message}\n")
    log_text.see(tk.END)
    log_text.config(state=tk.DISABLED)

# Effet hover sur les boutons
def on_enter(e):
    e.widget['background'] = '#e68099'  # Changement de couleur au survol

def on_leave(e):
    e.widget['background'] = '#ffc0cb'  # Couleur originale

# Créer une fenêtre Tkinter
root = tk.Tk()
root.title("ShipGrip")
root.geometry("600x500")
root.configure(bg="#d0e6f7")  # Bleu pastel pour le fond

# Charger et redimensionner le logo
logo_image = Image.open("logo.png")
logo_image = logo_image.resize((50, 50))  # Redimensionner l'image
logo = ImageTk.PhotoImage(logo_image)

# Afficher le logo et le titre
top_frame = tk.Frame(root, bg="#d0e6f7")
top_frame.pack(pady=10, padx=10, anchor='w')

logo_label = tk.Label(top_frame, image=logo, bg="#d0e6f7")
logo_label.pack(side='left', padx=5)

title_label = tk.Label(top_frame, text="ShipGrip", font=("Helvetica", 24, "bold"), bg="#d0e6f7", fg="black")
title_label.pack(side='left')

# Créer un bouton pour lancer la collecte des produits
btn_collecter = tk.Button(root, text="Collecter les produits", command=collecter_produits, bg="#ffc0cb", fg="black", font=("Helvetica", 14), padx=20, pady=10, relief="flat", bd=0)
btn_collecter.pack(pady=20)

# Appliquer l'effet hover sur le bouton
btn_collecter.bind("<Enter>", on_enter)
btn_collecter.bind("<Leave>", on_leave)

# Zone de texte pour afficher les logs
log_text = scrolledtext.ScrolledText(root, state=tk.DISABLED, height=10, width=70, bg="#ffc0cb", fg="black", font=("Helvetica", 12), relief="flat", bd=0)
log_text.pack(padx=20, pady=10)

# Lancer la boucle principale de l'interface
root.mainloop()